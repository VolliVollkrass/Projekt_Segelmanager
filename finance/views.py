import io
import re
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST

from boote.models import Boot
from toern.models import Teilnahme, Toern
from .models import Ausgabe, TopfAusgabe, TopfBeleg
from .utils import rate_kategorie

# Maximal pro Upload angenommene Belegfotos (gegen versehentliche Massen-Uploads).
MAX_BELEGE_PRO_UPLOAD = 20


def _gueltige_kategorie(raw):
    """Kategorie-Schlüssel validieren, sonst None."""
    gueltig = {k for k, _ in TopfAusgabe.KATEGORIE_CHOICES}
    return raw if raw in gueltig else None


def _speichere_belege(ausgabe, dateien, user):
    """Hochgeladene Belegfotos an eine Ausgabe hängen. Gibt Anzahl gespeicherter
    Belege zurück; ungültige Dateien werden übersprungen."""
    start = ausgabe.belege.count()
    gespeichert = 0
    for datei in dateien[:MAX_BELEGE_PRO_UPLOAD]:
        if not datei:
            continue
        try:
            TopfBeleg.objects.create(
                ausgabe=ausgabe,
                bild=datei,
                hochgeladen_von=user,
                order=start + gespeichert,
            )
            gespeichert += 1
        except Exception:
            # z.B. beschädigte Datei / kein Bild — überspringen
            continue
    return gespeichert


def _parse_betrag(raw):
    """Betrag aus dem Formular parsen (Komma oder Punkt), None wenn ungültig."""
    try:
        betrag = Decimal(str(raw).strip().replace(",", ".")).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
    except (InvalidOperation, AttributeError):
        return None
    if betrag <= 0:
        return None
    return betrag


def _ist_toern_skipper(user, toern):
    return Teilnahme.objects.filter(
        toern=toern, user=user, rolle__in=("skipper", "coskipper")
    ).exists()


# ───────────────────────── Bootskasse ─────────────────────────

@login_required
@require_POST
def ausgabe_erstellen(request, toern_id, boot_id):
    toern = get_object_or_404(Toern, id=toern_id)
    boot = get_object_or_404(Boot, id=boot_id, toern=toern)

    meine_teilnahme = Teilnahme.objects.filter(
        toern=toern, boot=boot, user=request.user, status="bestaetigt"
    ).first()
    if not meine_teilnahme:
        raise PermissionDenied

    kasse_url = f"{reverse('boot_dashboard', args=[toern.id])}?tab=kasse"

    beschreibung = request.POST.get("beschreibung", "").strip()
    betrag = _parse_betrag(request.POST.get("betrag"))
    bezahlt_von_id = request.POST.get("bezahlt_von")
    beteiligt_ids = request.POST.getlist("beteiligt")

    if not beschreibung or betrag is None:
        messages.error(request, "Bitte Beschreibung und einen gültigen Betrag angeben.")
        return redirect(kasse_url)

    zahler = Teilnahme.objects.filter(
        id=bezahlt_von_id, toern=toern, boot=boot, status="bestaetigt"
    ).first()
    if not zahler:
        messages.error(request, "Ungültiger Zahler.")
        return redirect(kasse_url)

    beteiligte = Teilnahme.objects.filter(
        id__in=beteiligt_ids, toern=toern, boot=boot, status="bestaetigt"
    )
    if not beteiligte.exists():
        messages.error(request, "Bitte mindestens eine beteiligte Person auswählen.")
        return redirect(kasse_url)

    ausgabe = Ausgabe.objects.create(
        boot=boot,
        toern=toern,
        beschreibung=beschreibung,
        betrag=betrag,
        bezahlt_von=zahler,
        erstellt_von=request.user,
    )
    ausgabe.beteiligt.set(beteiligte)

    messages.success(request, f"Ausgabe „{beschreibung}“ ({betrag} €) gespeichert.")
    return redirect(kasse_url)


@login_required
@require_POST
def ausgabe_bearbeiten(request, ausgabe_id):
    ausgabe = get_object_or_404(
        Ausgabe.objects.select_related("toern", "boot", "bezahlt_von__user"),
        id=ausgabe_id,
    )
    toern, boot = ausgabe.toern, ausgabe.boot

    darf_bearbeiten = (
        request.user == ausgabe.erstellt_von
        or request.user == ausgabe.bezahlt_von.user
        or request.user == toern.anbieter
        or _ist_toern_skipper(request.user, toern)
    )
    if not darf_bearbeiten:
        raise PermissionDenied

    kasse_url = f"{reverse('boot_dashboard', args=[toern.id])}?tab=kasse"

    beschreibung = request.POST.get("beschreibung", "").strip()
    betrag = _parse_betrag(request.POST.get("betrag"))
    bezahlt_von_id = request.POST.get("bezahlt_von")
    beteiligt_ids = request.POST.getlist("beteiligt")

    if not beschreibung or betrag is None:
        messages.error(request, "Bitte Beschreibung und einen gültigen Betrag angeben.")
        return redirect(kasse_url)

    zahler = Teilnahme.objects.filter(
        id=bezahlt_von_id, toern=toern, boot=boot, status="bestaetigt"
    ).first()
    if not zahler:
        messages.error(request, "Ungültiger Zahler.")
        return redirect(kasse_url)

    beteiligte = Teilnahme.objects.filter(
        id__in=beteiligt_ids, toern=toern, boot=boot, status="bestaetigt"
    )
    if not beteiligte.exists():
        messages.error(request, "Bitte mindestens eine beteiligte Person auswählen.")
        return redirect(kasse_url)

    ausgabe.beschreibung = beschreibung
    ausgabe.betrag = betrag
    ausgabe.bezahlt_von = zahler
    ausgabe.save()
    ausgabe.beteiligt.set(beteiligte)

    messages.success(request, f"Ausgabe „{beschreibung}“ aktualisiert.")
    return redirect(kasse_url)


@login_required
@require_POST
def ausgabe_loeschen(request, ausgabe_id):
    ausgabe = get_object_or_404(
        Ausgabe.objects.select_related("toern", "boot", "bezahlt_von__user"),
        id=ausgabe_id,
    )
    toern = ausgabe.toern

    darf_loeschen = (
        request.user == ausgabe.erstellt_von
        or request.user == ausgabe.bezahlt_von.user
        or request.user == toern.anbieter
        or _ist_toern_skipper(request.user, toern)
    )
    if not darf_loeschen:
        raise PermissionDenied

    ausgabe.delete()
    messages.success(request, "Ausgabe gelöscht.")
    return redirect(f"{reverse('boot_dashboard', args=[toern.id])}?tab=kasse")


# ───────────────────────── Skipper-Topf ─────────────────────────

@login_required
@require_POST
def topf_ausgabe_erstellen(request, toern_id):
    toern = get_object_or_404(Toern, id=toern_id)

    if request.user != toern.anbieter and not _ist_toern_skipper(request.user, toern):
        raise PermissionDenied

    kasse_url = f"{reverse('skipper_dashboard', args=[toern.id])}?tab=kasse"

    beschreibung = request.POST.get("beschreibung", "").strip()
    betrag = _parse_betrag(request.POST.get("betrag"))

    if not beschreibung or betrag is None:
        messages.error(request, "Bitte Beschreibung und einen gültigen Betrag angeben.")
        return redirect(kasse_url)

    # Kategorie: explizit gewählt, sonst automatisch aus der Beschreibung vorschlagen.
    kategorie = _gueltige_kategorie(request.POST.get("kategorie")) or rate_kategorie(beschreibung)

    ausgabe = TopfAusgabe.objects.create(
        toern=toern,
        erstellt_von=request.user,
        beschreibung=beschreibung,
        betrag=betrag,
        kategorie=kategorie,
    )

    anzahl = _speichere_belege(ausgabe, request.FILES.getlist("belege"), request.user)

    beleg_hinweis = f" · {anzahl} Beleg(e)" if anzahl else ""
    messages.success(
        request, f"Topf-Ausgabe „{beschreibung}“ ({betrag} €){beleg_hinweis} gespeichert."
    )
    return redirect(kasse_url)


@login_required
@require_POST
def topf_ausgabe_loeschen(request, ausgabe_id):
    ausgabe = get_object_or_404(
        TopfAusgabe.objects.select_related("toern"), id=ausgabe_id
    )
    toern = ausgabe.toern

    darf_loeschen = (
        request.user == ausgabe.erstellt_von
        or request.user == toern.anbieter
        or _ist_toern_skipper(request.user, toern)
    )
    if not darf_loeschen:
        raise PermissionDenied

    ausgabe.delete()
    messages.success(request, "Topf-Ausgabe gelöscht.")
    return redirect(f"{reverse('skipper_dashboard', args=[toern.id])}?tab=kasse")


def _darf_topf_verwalten(user, toern):
    """Skipper, Co-Skipper oder Anbieter dürfen den Topf verwalten & exportieren."""
    return user == toern.anbieter or _ist_toern_skipper(user, toern)


@login_required
@require_POST
def topf_ausgabe_bearbeiten(request, ausgabe_id):
    ausgabe = get_object_or_404(
        TopfAusgabe.objects.select_related("toern"), id=ausgabe_id
    )
    toern = ausgabe.toern

    if not (request.user == ausgabe.erstellt_von or _darf_topf_verwalten(request.user, toern)):
        raise PermissionDenied

    kasse_url = f"{reverse('skipper_dashboard', args=[toern.id])}?tab=kasse"

    beschreibung = request.POST.get("beschreibung", "").strip()
    betrag = _parse_betrag(request.POST.get("betrag"))
    if not beschreibung or betrag is None:
        messages.error(request, "Bitte Beschreibung und einen gültigen Betrag angeben.")
        return redirect(kasse_url)

    ausgabe.beschreibung = beschreibung
    ausgabe.betrag = betrag
    ausgabe.kategorie = _gueltige_kategorie(request.POST.get("kategorie")) or ausgabe.kategorie
    ausgabe.save()

    # Optional beim Bearbeiten direkt weitere Belege mitschicken.
    _speichere_belege(ausgabe, request.FILES.getlist("belege"), request.user)

    messages.success(request, f"Topf-Ausgabe „{beschreibung}“ aktualisiert.")
    return redirect(kasse_url)


@login_required
@require_POST
def topf_beleg_add(request, ausgabe_id):
    ausgabe = get_object_or_404(
        TopfAusgabe.objects.select_related("toern"), id=ausgabe_id
    )
    toern = ausgabe.toern

    if not (request.user == ausgabe.erstellt_von or _darf_topf_verwalten(request.user, toern)):
        raise PermissionDenied

    kasse_url = f"{reverse('skipper_dashboard', args=[toern.id])}?tab=kasse"
    anzahl = _speichere_belege(ausgabe, request.FILES.getlist("belege"), request.user)

    if anzahl:
        messages.success(request, f"{anzahl} Beleg(e) hinzugefügt.")
    else:
        messages.error(request, "Keine gültigen Belegfotos empfangen.")
    return redirect(kasse_url)


@login_required
@require_POST
def topf_beleg_loeschen(request, beleg_id):
    beleg = get_object_or_404(
        TopfBeleg.objects.select_related("ausgabe__toern"), id=beleg_id
    )
    toern = beleg.ausgabe.toern

    if not (request.user == beleg.ausgabe.erstellt_von or _darf_topf_verwalten(request.user, toern)):
        raise PermissionDenied

    beleg.delete()
    messages.success(request, "Beleg gelöscht.")
    return redirect(f"{reverse('skipper_dashboard', args=[toern.id])}?tab=kasse")


# ───────────────────────── Abrechnung: gemeinsame Daten ─────────────────────────

def _abrechnung_struktur(toern):
    """Ausgaben des Topfs nach Kategorie gruppiert, mit fortlaufenden Beleg-Nummern.

    Dieselbe Reihenfolge/Nummerierung wird von PDF und Excel genutzt, damit die
    Beleg-Nr. in der Excel-Liste auf die richtige Seite im PDF verweist.
    """
    reihenfolge = [k for k, _ in TopfAusgabe.KATEGORIE_CHOICES]
    labels = dict(TopfAusgabe.KATEGORIE_CHOICES)

    ausgaben = list(
        TopfAusgabe.objects.filter(toern=toern)
        .select_related("erstellt_von")
        .prefetch_related("belege")
    )
    # Nach Kategorie-Reihenfolge, dann chronologisch (älteste zuerst).
    ausgaben.sort(key=lambda a: (reihenfolge.index(a.kategorie) if a.kategorie in reihenfolge else 999, a.created_at))

    gruppen = []
    beleg_nr = 0
    for key in reihenfolge:
        gruppe_ausgaben = [a for a in ausgaben if a.kategorie == key]
        if not gruppe_ausgaben:
            continue
        eintraege = []
        summe = Decimal("0")
        for a in gruppe_ausgaben:
            summe += a.betrag
            belege = []
            for b in a.belege.all():
                beleg_nr += 1
                belege.append({"beleg": b, "nummer": f"B-{beleg_nr:03d}"})
            eintraege.append({"ausgabe": a, "belege": belege})
        gruppen.append({
            "key": key,
            "label": labels[key],
            "summe": summe,
            "eintraege": eintraege,
        })

    gesamt = sum((g["summe"] for g in gruppen), Decimal("0"))
    return gruppen, gesamt


def _dateiname(toern, endung):
    basis = re.sub(r"[^\w\-]", "_", toern.titel or "Toern")[:50]
    return f"Abrechnung_{basis}.{endung}"


@login_required
def topf_belege_pdf(request, toern_id):
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.platypus import Image as RLImage
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from PIL import Image as PILImage

    toern = get_object_or_404(Toern, id=toern_id)
    if not _darf_topf_verwalten(request.user, toern):
        raise PermissionDenied

    gruppen, gesamt = _abrechnung_struktur(toern)

    buffer = io.BytesIO()
    MAR = 18 * mm
    USABLE_W = A4[0] - 2 * MAR
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=MAR, rightMargin=MAR, topMargin=MAR, bottomMargin=MAR,
        title=f"Belegsammlung – {toern.titel}",
    )

    PRIMARY = colors.HexColor("#1e3a5f")
    TEAL = colors.HexColor("#0D9488")

    title_s = ParagraphStyle("t", fontSize=20, leading=24, textColor=PRIMARY, fontName="Helvetica-Bold")
    sub_s = ParagraphStyle("s", fontSize=10, leading=14, textColor=colors.HexColor("#64748b"))
    cat_s = ParagraphStyle("c", fontSize=13, leading=17, textColor=TEAL, fontName="Helvetica-Bold", spaceBefore=6*mm, spaceAfter=1*mm)
    exp_s = ParagraphStyle("e", fontSize=10, leading=14, fontName="Helvetica-Bold", textColor=colors.black)
    meta_s = ParagraphStyle("m", fontSize=8, leading=11, textColor=colors.HexColor("#64748b"))
    num_s = ParagraphStyle("n", fontSize=8, leading=11, fontName="Helvetica-Bold", textColor=TEAL)
    note_s = ParagraphStyle("no", fontSize=9, leading=12, fontName="Helvetica-Oblique", textColor=colors.HexColor("#b45309"))

    elements = []
    elements.append(Paragraph("Belegsammlung – Skipper-Topf", title_s))
    elements.append(Paragraph(
        f"{toern.titel} · Stand {timezone.localdate().strftime('%d.%m.%Y')} · "
        f"Budget {toern.skipper_budget:.2f} € · Ausgegeben {gesamt:.2f} € · "
        f"Verbleibend {(toern.skipper_budget - gesamt):.2f} €",
        sub_s,
    ))
    elements.append(HRFlowable(width=USABLE_W, thickness=0.5, color=TEAL, spaceBefore=2*mm, spaceAfter=1*mm))

    if not gruppen:
        elements.append(Spacer(1, 6*mm))
        elements.append(Paragraph("Noch keine Ausgaben erfasst.", meta_s))

    MAX_IMG_H = 150 * mm
    for g in gruppen:
        elements.append(Paragraph(f"{g['label']}  ·  {g['summe']:.2f} €", cat_s))
        elements.append(HRFlowable(width=USABLE_W, thickness=0.3, color=colors.HexColor("#e2e8f0")))
        for eintrag in g["eintraege"]:
            a = eintrag["ausgabe"]
            erfasser = f"{a.erstellt_von.first_name} {a.erstellt_von.last_name}".strip() if a.erstellt_von else "—"
            elements.append(Spacer(1, 2*mm))
            elements.append(Paragraph(f"{a.beschreibung} — {a.betrag:.2f} €", exp_s))
            elements.append(Paragraph(f"{a.created_at.strftime('%d.%m.%Y')} · {erfasser}", meta_s))
            if not eintrag["belege"]:
                elements.append(Paragraph("⚠ Kein Beleg vorhanden", note_s))
                continue
            for bl in eintrag["belege"]:
                try:
                    path = bl["beleg"].bild.path
                    with PILImage.open(path) as im:
                        iw, ih = im.size
                    ratio = ih / iw if iw else 1
                    w = USABLE_W
                    h = w * ratio
                    if h > MAX_IMG_H:
                        h = MAX_IMG_H
                        w = h / ratio
                    elements.append(Spacer(1, 1.5*mm))
                    elements.append(Paragraph(bl["nummer"], num_s))
                    elements.append(RLImage(path, width=w, height=h))
                except Exception:
                    elements.append(Paragraph(f"{bl['nummer']} · Beleg konnte nicht geladen werden", meta_s))

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{_dateiname(toern, "pdf")}"'
    return response


@login_required
def topf_abrechnung_xlsx(request, toern_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    toern = get_object_or_404(Toern, id=toern_id)
    if not _darf_topf_verwalten(request.user, toern):
        raise PermissionDenied

    gruppen, gesamt = _abrechnung_struktur(toern)

    wb = Workbook()
    ws = wb.active
    ws.title = "Abrechnung"

    PRIMARY = "1E3A5F"
    TEAL = "0D9488"
    LIGHT = "E2E8F0"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=PRIMARY)
    cat_font = Font(bold=True, color="FFFFFF")
    cat_fill = PatternFill("solid", fgColor=TEAL)
    bold = Font(bold=True)
    thin = Side(style="thin", color=LIGHT)
    border = Border(bottom=thin)
    euro = '#,##0.00\\ €'

    # Titel
    ws.append([f"Abrechnung Skipper-Topf – {toern.titel}"])
    ws["A1"].font = Font(bold=True, size=14, color=PRIMARY)
    ws.append([f"Stand {timezone.localdate().strftime('%d.%m.%Y')}"])
    ws["A2"].font = Font(italic=True, color="64748B")
    ws.append([])

    spalten = ["Datum", "Kategorie", "Beschreibung", "Betrag", "Erfasst von", "Belege (siehe PDF)"]
    header_row = ws.max_row + 1
    ws.append(spalten)
    for col in range(1, len(spalten) + 1):
        c = ws.cell(row=header_row, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    for g in gruppen:
        # Kategorie-Zeile
        r = ws.max_row + 1
        ws.append([g["label"], "", "", g["summe"], "", ""])
        for col in range(1, 7):
            ws.cell(row=r, column=col).fill = cat_fill
            ws.cell(row=r, column=col).font = cat_font
        ws.cell(row=r, column=4).number_format = euro
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="right")

        for eintrag in g["eintraege"]:
            a = eintrag["ausgabe"]
            erfasser = f"{a.erstellt_von.first_name} {a.erstellt_von.last_name}".strip() if a.erstellt_von else ""
            belege_ref = ", ".join(bl["nummer"] for bl in eintrag["belege"]) or "— kein Beleg —"
            rr = ws.max_row + 1
            ws.append([
                a.created_at.strftime("%d.%m.%Y"),
                g["label"],
                a.beschreibung,
                float(a.betrag),
                erfasser,
                belege_ref,
            ])
            ws.cell(row=rr, column=4).number_format = euro
            ws.cell(row=rr, column=4).alignment = Alignment(horizontal="right")
            for col in range(1, 7):
                ws.cell(row=rr, column=col).border = border

    # Summenblock
    ws.append([])
    def summenzeile(label, wert, fett=True):
        r = ws.max_row + 1
        ws.append(["", "", label, float(wert), "", ""])
        ws.cell(row=r, column=3).font = bold if fett else Font()
        c = ws.cell(row=r, column=4)
        c.number_format = euro
        c.font = bold if fett else Font()
        c.alignment = Alignment(horizontal="right")
        return r

    for g in gruppen:
        summenzeile(g["label"], g["summe"], fett=False)
    ws.append([])
    summenzeile("Gesamt ausgegeben", gesamt)
    summenzeile("Budget", toern.skipper_budget)
    rest_row = summenzeile("Verbleibend", toern.skipper_budget - gesamt)
    if toern.skipper_budget - gesamt < 0:
        ws.cell(row=rest_row, column=4).font = Font(bold=True, color="DC2626")

    breiten = [12, 30, 40, 14, 22, 24]
    for i, w in enumerate(breiten, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{_dateiname(toern, "xlsx")}"'
    return response
